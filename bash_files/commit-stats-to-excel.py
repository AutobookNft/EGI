#!/usr/bin/env python3
"""
📊 EGI Commit Statistics to Excel Converter
===========================================

Converte le statistiche dei commit in un file Excel ben formattato.
Analizza i commit per settimana dal 19 agosto 2025 (introduzione TAG system).

@author: AI Partner OS2.0-Compliant for Fabio Cherici
@version: 1.0.0 (FlorenceEGI MVP - Excel Export System)
@os2-pillars: Explicit,Coherent,Simple,Secure
"""

import subprocess
import pandas as pd
import re
from datetime import datetime, timedelta
import sys
import os
from pathlib import Path

class EGICommitStatsExporter:
    def __init__(self):
        self.git_repo_path = Path(__file__).parent.parent
        self.output_file = self.git_repo_path / "commit_statistics.xlsx"
        self.tag_patterns = {
            'FEAT': r'\[FEAT\]',
            'FIX': r'\[FIX\]', 
            'REFACTOR': r'\[REFACTOR\]',
            'DOC': r'\[DOC\]',
            'TEST': r'\[TEST\]',
            'CHORE': r'\[CHORE\]'
        }
        
    def run_git_command(self, command):
        """Esegue un comando git e ritorna l'output"""
        try:
            result = subprocess.run(
                command, 
                shell=True, 
                cwd=self.git_repo_path,
                capture_output=True, 
                text=True, 
                check=True
            )
            return result.stdout.strip()
        except subprocess.CalledProcessError as e:
            print(f"❌ Errore git command: {e}")
            return ""
    
    def get_commits_for_period(self, start_date, end_date):
        """Ottieni tutti i commit per un periodo specifico"""
        cmd = f'git log --oneline --since="{start_date}" --until="{end_date} 23:59:59"'
        output = self.run_git_command(cmd)
        
        if not output:
            return []
            
        commits = []
        for line in output.split('\n'):
            if line.strip():
                commits.append(line.strip())
        
        return commits
    
    def analyze_commits(self, commits):
        """Analizza i commit e categorizza per TAG"""
        stats = {
            'total_commits': len(commits),
            'tagged_commits': 0,
            'untagged_commits': 0,
            'tags': {tag: 0 for tag in self.tag_patterns.keys()}
        }
        
        for commit in commits:
            has_tag = False
            for tag, pattern in self.tag_patterns.items():
                if re.search(pattern, commit):
                    stats['tags'][tag] += 1
                    has_tag = True
                    break
            
            if has_tag:
                stats['tagged_commits'] += 1
            else:
                stats['untagged_commits'] += 1
        
        # Calcola percentuali
        if stats['total_commits'] > 0:
            stats['tag_coverage'] = round((stats['tagged_commits'] / stats['total_commits']) * 100, 1)
        else:
            stats['tag_coverage'] = 0
            
        return stats
    
    def get_daily_commits(self, start_date, end_date):
        """Ottieni commit giornalieri per il periodo"""
        daily_stats = []
        current_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_dt = datetime.strptime(end_date, '%Y-%m-%d')
        
        while current_date <= end_dt:
            date_str = current_date.strftime('%Y-%m-%d')
            next_date = current_date + timedelta(days=1)
            next_date_str = next_date.strftime('%Y-%m-%d')
            
            cmd = f'git log --oneline --since="{date_str}" --until="{next_date_str}"'
            output = self.run_git_command(cmd)
            
            commits = [line.strip() for line in output.split('\n') if line.strip()]
            commit_count = len(commits)
            
            daily_stats.append({
                'date': date_str,
                'day_name': current_date.strftime('%A'),
                'commits': commit_count
            })
            
            current_date += timedelta(days=1)
            
        return daily_stats
    
    def get_testing_time_data(self, start_date, end_date):
        """Estrae dati di testing time dal log per il periodo specificato"""
        testing_log_path = self.git_repo_path / "storage" / "logs" / "testing_time.log"
        
        if not testing_log_path.exists():
            return {
                'total_minutes': 0,
                'sessions_count': 0,
                'avg_session_minutes': 0,
                'daily_breakdown': {}
            }
        
        import json
        from datetime import datetime, timedelta
        
        # Parse date range
        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
        end_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
        
        total_minutes = 0
        sessions_count = 0
        daily_breakdown = {}
        
        try:
            with open(testing_log_path, 'r') as f:
                for line in f:
                    try:
                        data = json.loads(line.strip())
                        if data['action'] == 'TESTING_END':
                            timestamp = datetime.fromisoformat(data['timestamp'].replace('Z', '+00:00')).replace(tzinfo=None)
                            
                            if start_dt <= timestamp < end_dt:
                                # Usa il valore assoluto per gestire durate negative dei dati storici
                                duration = abs(data.get('duration', 0))
                                date_key = timestamp.strftime('%Y-%m-%d')
                                
                                total_minutes += duration
                                sessions_count += 1
                                
                                if date_key not in daily_breakdown:
                                    daily_breakdown[date_key] = {'minutes': 0, 'sessions': 0}
                                daily_breakdown[date_key]['minutes'] += duration
                                daily_breakdown[date_key]['sessions'] += 1
                    except:
                        continue
        except:
            pass
        
        avg_session_minutes = total_minutes / sessions_count if sessions_count > 0 else 0
        
        return {
            'total_minutes': total_minutes,
            'sessions_count': sessions_count,
            'avg_session_minutes': round(avg_session_minutes, 1),
            'daily_breakdown': daily_breakdown
        }
    
    def generate_weekly_data(self):
        """Genera dati settimanali dal 19 agosto 2025 con testing time"""
        weeks = [
            {
                'name': 'Settimana 1',
                'period': '19-25 Agosto 2025',
                'start_date': '2025-08-19',
                'end_date': '2025-08-25',
                'description': 'Introduzione TAG system'
            },
            {
                'name': 'Settimana 2', 
                'period': '26 Ago - 1 Set 2025',
                'start_date': '2025-08-26',
                'end_date': '2025-09-01',
                'description': 'Stabilizzazione'
            },
            {
                'name': 'Settimana 3',
                'period': '2-8 Settembre 2025',
                'start_date': '2025-09-02',
                'end_date': '2025-09-08',
                'description': 'Consolidamento (in corso)'
            }
        ]
        
        weekly_data = []
        all_daily_data = []
        testing_summary = []
        
        for week in weeks:
            # Analisi settimanale commit
            commits = self.get_commits_for_period(week['start_date'], week['end_date'])
            stats = self.analyze_commits(commits)
            
            # Analisi testing time
            testing_data = self.get_testing_time_data(week['start_date'], week['end_date'])
            
            # Stima tempo coding (22 min per commit)
            estimated_coding_minutes = stats['total_commits'] * 22
            total_productive_minutes = testing_data['total_minutes'] + estimated_coding_minutes
            
            weekly_data.append({
                'Settimana': week['name'],
                'Periodo': week['period'],
                'Descrizione': week['description'],
                'Commit Totali': stats['total_commits'],
                'Commit con TAG': stats['tagged_commits'],
                'Commit senza TAG': stats['untagged_commits'],
                'Copertura TAG %': stats['tag_coverage'],
                'FEAT': stats['tags']['FEAT'],
                'FIX': stats['tags']['FIX'],
                'REFACTOR': stats['tags']['REFACTOR'],
                'DOC': stats['tags']['DOC'],
                'TEST': stats['tags']['TEST'],
                'CHORE': stats['tags']['CHORE'],
                'TAG Dominante': max(stats['tags'], key=stats['tags'].get) if any(stats['tags'].values()) else 'Nessuno',
                'Testing Minutes': testing_data['total_minutes'],
                'Testing Sessions': testing_data['sessions_count'],
                'Avg Session (min)': testing_data['avg_session_minutes'],
                'Coding Minutes (est)': estimated_coding_minutes,
                'Total Productive Minutes': total_productive_minutes,
                'Testing %': round((testing_data['total_minutes'] / total_productive_minutes) * 100, 1) if total_productive_minutes > 0 else 0
            })
            
            # Dati giornalieri con testing
            daily_data = self.get_daily_commits(week['start_date'], week['end_date'])
            for day in daily_data:
                day['settimana'] = week['name']
                date_str = day['date']
                
                # Aggiungi dati testing per questo giorno
                if date_str in testing_data['daily_breakdown']:
                    day['testing_minutes'] = testing_data['daily_breakdown'][date_str]['minutes']
                    day['testing_sessions'] = testing_data['daily_breakdown'][date_str]['sessions']
                else:
                    day['testing_minutes'] = 0
                    day['testing_sessions'] = 0
                
                # Calcola tempo produttivo totale
                day['coding_minutes_est'] = day['commits'] * 22
                day['total_productive_minutes'] = day['testing_minutes'] + day['coding_minutes_est']
                
                all_daily_data.append(day)
            
            # Summary testing per settimana
            testing_summary.append({
                'Settimana': week['name'],
                'Periodo': week['period'],
                'Testing Totale (h)': round(testing_data['total_minutes'] / 60, 1),
                'Sessioni Totali': testing_data['sessions_count'],
                'Media Sessione (min)': testing_data['avg_session_minutes'],
                'Coding Stimato (h)': round(estimated_coding_minutes / 60, 1),
                'Tempo Produttivo (h)': round(total_productive_minutes / 60, 1),
                'Rapporto Testing/Coding': f"{round((testing_data['total_minutes'] / estimated_coding_minutes) * 100, 1)}%" if estimated_coding_minutes > 0 else "N/A"
            })
        
        return weekly_data, all_daily_data, testing_summary
    
    def create_excel_file(self):
        """Crea il file Excel con tutti i dati inclusi testing time"""
        print("📊 Generazione statistiche commit e testing per Excel...")
        
        # Genera dati
        weekly_data, daily_data, testing_summary = self.generate_weekly_data()
        
        # Crea DataFrames
        df_weekly = pd.DataFrame(weekly_data)
        df_daily = pd.DataFrame(daily_data)
        df_testing = pd.DataFrame(testing_summary)
        
        # Dati di riepilogo
        total_commits = sum(week['Commit Totali'] for week in weekly_data)
        total_tagged = sum(week['Commit con TAG'] for week in weekly_data)
        avg_coverage = round(sum(week['Copertura TAG %'] for week in weekly_data) / len(weekly_data), 1)
        total_testing_minutes = sum(week['Testing Minutes'] for week in weekly_data)
        total_coding_minutes = sum(week['Coding Minutes (est)'] for week in weekly_data)
        
        summary_data = [{
            'Metrica': 'Commit Totali',
            'Valore': total_commits,
            'Note': 'Dal 19 agosto 2025'
        }, {
            'Metrica': 'Commit con TAG',
            'Valore': total_tagged,
            'Note': f'{round((total_tagged/total_commits)*100, 1)}% del totale'
        }, {
            'Metrica': 'Giorni con TAG System',
            'Valore': (datetime.now() - datetime(2025, 8, 19)).days + 1,
            'Note': 'Dal 19 agosto 2025'
        }, {
            'Metrica': 'Copertura TAG Media',
            'Valore': f'{avg_coverage}%',
            'Note': 'Media delle 3 settimane'
        }, {
            'Metrica': 'Testing Time Totale',
            'Valore': f'{round(total_testing_minutes/60, 1)}h',
            'Note': f'{total_testing_minutes} minuti'
        }, {
            'Metrica': 'Coding Time Stimato',
            'Valore': f'{round(total_coding_minutes/60, 1)}h',
            'Note': '22 min per commit'
        }, {
            'Metrica': 'Rapporto Testing/Coding',
            'Valore': f'{round((total_testing_minutes/total_coding_minutes)*100, 1)}%' if total_coding_minutes > 0 else 'N/A',
            'Note': 'Testing vs sviluppo'
        }]
        
        df_summary = pd.DataFrame(summary_data)
        
        # Scrivi file Excel
        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            # Sheet 1: Riepilogo
            df_summary.to_excel(writer, sheet_name='Riepilogo', index=False)
            
            # Sheet 2: Dati Settimanali  
            df_weekly.to_excel(writer, sheet_name='Statistiche Settimanali', index=False)
            
            # Sheet 3: Testing Summary
            df_testing.to_excel(writer, sheet_name='Testing Time Analysis', index=False)
            
            # Sheet 4: Dati Giornalieri
            df_daily.to_excel(writer, sheet_name='Commit Giornalieri', index=False)
            
            # Formattazione
            self.format_excel_sheets(writer)
        
        print(f"✅ File Excel creato: {self.output_file}")
        print(f"📁 Percorso completo: {self.output_file.absolute()}")
        print(f"📊 Testing time totale: {round(total_testing_minutes/60, 1)}h")
        print(f"💻 Coding time stimato: {round(total_coding_minutes/60, 1)}h")
        
        return str(self.output_file.absolute())
    
    def format_excel_sheets(self, writer):
        """Formatta i fogli Excel"""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Stili
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        # Formatta ogni sheet
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            
            # Header styling
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

def main():
    """Funzione principale"""
    print("🚀 EGI Commit Statistics Excel Exporter")
    print("=" * 50)
    
    exporter = EGICommitStatsExporter()
    
    try:
        output_path = exporter.create_excel_file()
        print(f"\n🎉 Export completato con successo!")
        print(f"📊 File salvato in: {output_path}")
        
        # Verifica se il file esiste
        if os.path.exists(output_path):
            file_size = os.path.getsize(output_path)
            print(f"📁 Dimensione file: {file_size} bytes")
        
    except Exception as e:
        print(f"❌ Errore durante l'export: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
